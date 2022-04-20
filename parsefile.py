import openpyxl
import os
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from apps.resource.models import ResourceItem
from apps.companies_11736.models import Company, ChangeLog
from apps.companies_11736.helpers import get_managers_json, get_founders_json, get_beneficiaries_json, \
    get_activities_json, log_changes, split_list, Timer


def get_all_parsed_data(active_sheet):
    # find headers
    headers = {
        "idno": None,
        "register_date": None,
        "name": None,
        "organization_form": None,
        "address": None,
        "cuatm": None,
        "managers": None,
        "founders": None,
        "beneficiaries": None,
        "state": None,
        "liquidation_date": None,
        "licensed_activities": None,
        "unlicensed_activities": None
    }
    rows = active_sheet.max_row + 1
    # to get last column in loop
    columns = active_sheet.max_column + 1
    # check if we have title in first row
    if active_sheet.cell(1, 1) == "":
        i = 2
        start_row = 3
    else:
        i = 1
        start_row = 2

    for j in range(1, columns):
        cell_value = str(active_sheet.cell(row=i, column=j).value)
        if cell_value.startswith('IDNO'):
            headers['idno'] = j - 1
        if cell_value.startswith('Data înregistrării'):
            headers['register_date'] = j - 1
        if cell_value.startswith('Denumirea completă'):
            headers['name'] = j - 1
        if cell_value.startswith('Forma org'):
            headers['organization_form'] = j - 1
        if cell_value.startswith('Adresa'):
            headers['address'] = j - 1
        if cell_value.startswith('Codul'):
            headers['cuatm'] = j - 1
        if cell_value.startswith('Lista conducătorilor'):
            headers['managers'] = j - 1
        if cell_value.startswith('Lista fondatorilor'):
            headers['founders'] = j - 1
        if cell_value.startswith('Lista beneficiarilor'):
            headers['beneficiaries'] = j - 1
        if cell_value.startswith('Stat'):
            headers['state'] = j - 1
        if cell_value.startswith('Data lichidării'):
            headers['liquidation_date'] = j - 1
        if cell_value.startswith('Genuri de activitate licentiate'):
            headers['licensed_activities'] = j - 1
        if cell_value.startswith('Genuri de activitate nelicentiate'):
            headers['unlicensed_activities'] = j - 1

    # -------------------------------------------------

    # get data

    data = []

    for i in range(start_row, rows):
        values = [active_sheet.cell(row=i, column=j).value for j in range(1, columns)]

        # this values always exist in excel file
        d = {
            "idno": values[headers['idno']],
            "register_date": values[headers['register_date']],
            "name": values[headers['name']],
            "organization_form": values[headers['organization_form']],
            "address": values[headers['address']],
            "managers": get_managers_json(
                str(values[headers['managers']])
            )
        }
        # this values need to be checked
        if headers['cuatm']:
            d['cuatm'] = values[headers['cuatm']]
        else:
            d['cuatm'] = None

        if headers['founders']:
            d['founders'] = get_founders_json(
                str(values[headers['founders']])
            )
        else:
            d['founders'] = None

        if headers['beneficiaries']:
            d['beneficiaries'] = get_beneficiaries_json(
                str(values[headers['beneficiaries']])
            )
        else:
            d['beneficiaries'] = None

        if headers['state']:
            d['state'] = values[headers['state']]
        else:
            d['state'] = None

        if headers['liquidation_date']:
            d['liquidation_date'] = values[headers['liquidation_date']]
        else:
            d['liquidation_date'] = None

        if headers['licensed_activities']:
            d['licensed_activities'] = get_activities_json(
                str(values[headers['licensed_activities']])
            )
        else:
            d['licensed_activities'] = None

        if headers['unlicensed_activities']:
            d['unlicensed_activities'] = get_activities_json(
                str(values[headers['unlicensed_activities']])
            )
        else:
            d['unlicensed_activities'] = None

        # append dict with data in list
        data.append(d)

    return data


def parse_docs():
    files = ResourceItem.objects.filter(parsed=False).order_by('created_at')

    for file in files:
        # open excel file
        excel_file = openpyxl.load_workbook(file.file)
        # parse only main sheet that contains the data
        active_sheet = excel_file.active

        data = get_all_parsed_data(active_sheet)

        changes = []
        to_create = []
        to_update = []

        all_companies = Company.objects.all()

        for item in data:
            company = all_companies.filter(idno=item['idno'], name=item['name'])
            # if company exists then need to be updated
            if company.exists():
                # log changes
                change = log_changes(item, company.first())
                if change:
                    changes.append(
                        ChangeLog(
                            company_id=company.first().id,
                            change=change,
                        )
                    )
                # append to list for updating
                to_update.append(
                    Company(
                        id=company.first().id,
                        **item,
                    )
                )
            # else need to be created
            else:
                # append to list for creating
                to_create.append(
                    Company(
                        **item,
                    )
                )

        ChangeLog.objects.bulk_create(changes)
        Company.objects.bulk_create(to_create, ignore_conflicts=True)

        divided_list = split_list(alist=to_update, wanted_parts=20)
        for lists in divided_list:
            Company.objects.bulk_update(lists,
                                        ['name', 'organization_form', 'address', 'cuatm', 'managers', 'founders',
                                         'beneficiaries', 'state', 'liquidation_date', 'licensed_activities',
                                         'unlicensed_activities'], batch_size=len(lists))

        file.parsed = True
        file.save()

        break


parse_docs()
